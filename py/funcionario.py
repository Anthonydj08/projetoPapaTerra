from openpyxl import Workbook
from openpyxl import load_workbook

def criarBancoDeDados():
  arquivo_excel = Workbook()
  planilha1 = arquivo_excel.active
  planilha1.title = "Voos"
  planilha1 = arquivo_excel.create_sheet("Passageiros")
  
  planilha1 = arquivo_excel.get_sheet_by_name("Voos")

  planilha1["A1"] = "codVoo"
  planilha1["B1"] = "aeroportoPartida"
  planilha1["C1"] = "horarioSaida"
  planilha1["D1"] = "aeroportoDestino"
  planilha1["E1"] = "horarioChegada"
  planilha1["F1"] = "quantidadeAssentos"

  planilha1 = arquivo_excel.get_sheet_by_name("Passageiros")

  planilha1["A1"] = "codPassageiro"
  planilha1["B1"] = "nome"
  planilha1["C1"] = "cpf"
  planilha1["D1"] = "telefone"
  planilha1["E1"] = "email"
  
  arquivo_excel.save("BD.xlsx")
  
  
def cadastrarVoo2(aeroPartida, horarioSaida, aeroDestino, horarioChegada, quantidadeAssentos):
  wb = load_workbook('BD.xlsx')
  sheet = wb["Voos"]

  codVoo = sheet.max_row
  
  sheet.append((codVoo, aeroPartida, horarioSaida, aeroDestino, horarioChegada, quantidadeAssentos))
  
  wb.create_sheet('Voo'+str(codVoo))
  BD = wb.get_sheet_by_name('Voo'+str(codVoo))
  BD['A1'] = 'codVoo'
  BD['B1'] = 'codPassageiro'
  BD['C1'] = 'numeroAssento'
  BD['D1'] = 'classePassageiro'
  BD['E1'] = 'valorGasto'

  wb.save("BD.xlsx")

def cadastrarVoo():
  wb = load_workbook('BD.xlsx')
  sheet = wb["Voos"]

  codVoo = sheet.max_row
  aeroPartida = input('Aeroporto de partida ')
  horarioSaida = input('Horário de Saida ')
  aeroDestino = input('Aeroporto de destino ')
  horarioChegada = input('Horário de chegada ')
  quantidadeAssentos = input('Quantidade de assentos ')
  
  sheet.append((codVoo, aeroPartida, horarioSaida, aeroDestino, horarioChegada, quantidadeAssentos))
  
  wb.create_sheet('Voo'+str(codVoo))
  BD = wb.get_sheet_by_name('Voo'+str(codVoo))
  BD['A1'] = 'codVoo'
  BD['B1'] = 'codPassageiro'
  BD['C1'] = 'numeroAssento'
  BD['D1'] = 'classePassageiro'
  BD['E1'] = 'valorGasto'

  wb.save("BD.xlsx")

def cadastrarPassageiro():
  wb = load_workbook('BD.xlsx')
  sheet = wb["Passageiros"]

  codPassageiro = sheet.max_row
  nome = input('Nome do passageiro ')
  cpf = input('CPF do passageiro ')
  telefone = input('Telefone do passageiro ')
  email = input('Email do passageiro ')
    
  sheet.append((codPassageiro, nome, cpf, telefone, email))

  wb.save("BD.xlsx")

def inserirPassageiro():
  wb = load_workbook('BD.xlsx')

  sheet = wb["Voos"]
  max_linha = sheet.max_row + 1
  max_coluna = sheet.max_column + 1
  matrizV = []

  for i in range(1, max_linha):
    linhaV = []
    for j in range(1, max_coluna):
      linhaV.append(sheet.cell(column=j, row=i).value)
    matrizV.append(linhaV)
    
  for i in range(0, max_linha - 1):
    for j in range(0, max_coluna - 1):
      print(f'[{matrizV[i][j]:^20}]', end=' ')
    print()

  codVoo = input('Código do Voo ')
  quantidadeAssentos = sheet['F'+str(int(codVoo) + 1)].value

  sheet = wb["Passageiros"]
  max_linha = sheet.max_row + 1
  max_coluna = sheet.max_column + 1
  matrizP = []

  for i in range(1, max_linha):
    linhaP = []
    for j in range(1, max_coluna):
      linhaP.append(sheet.cell(column=j, row=i).value)
    matrizP.append(linhaP)
  for i in range(0, max_linha - 1):
    for j in range(0, max_coluna - 1):
      print(f'[{matrizP[i][j]:^20}]', end=' ')
    print()

  codPassageiro = input('Código do Passageiro ')

  sheet = wb['Voo'+str(codVoo)]

  valida = False
  while valida == False:
    numeroAssento = int(input('Número do assento '))
    cont = 1
    max_linhas = sheet.max_row
    for i in range(1, max_linhas + 1):
      if sheet['C'+str(int(i))].value == numeroAssento:
        break
      elif sheet['C'+str(int(i))].value != numeroAssento and cont == max_linhas:
        if numeroAssento > 0 and numeroAssento <= int(quantidadeAssentos):
          valida = True
      cont = cont + 1
          
  valida = False
  while valida == False:
    classePassageiro = input('Classe do passageiro ')
    if(classePassageiro == 'economica' or classePassageiro == 'executiva'):
      valida = True

  valorGasto = 0

  sheet.append((codVoo, codPassageiro, numeroAssento, classePassageiro, valorGasto))

  wb.save("BD.xlsx")

def listarVoos():
  wb = load_workbook('BD.xlsx')
  sheet = wb["Voos"]
  max_linha = sheet.max_row + 1
  max_coluna = sheet.max_column + 1
  matrizV = []
  voos = []
  for i in range(1, max_linha):
    linhaV = []
    for j in range(1, max_coluna):
      linhaV.append(sheet.cell(column=j, row=i).value)
    voos.insert(max_linha, linhaV)
  return voos

def func():
  sair = False
  while (sair == False):
    op = int(input('Digite:\n1 - Cadastrar Voo \n2 - Cadastrar Passageiro\n3 - Inserir Passageiro em um Voo\n4 - Verificar Lucro\n5 - Informações sobre Voos\n0 - Sair\n'))
    if op == 1:
      cadastrarVoo()
    if op == 2:
      cadastrarPassageiro()
    if op == 3:
      inserirPassageiro()
    # if op == 4:
    #   verificarLucro()
    # if op == 5:
    #   infoVoo()
    elif op == 0:
      sair = True